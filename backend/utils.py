import csv
import io
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def rows_to_csv(rows):
    if not rows:
        return ""
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=list(rows[0].keys()))
    writer.writeheader()
    writer.writerows(rows)
    return buffer.getvalue()


def rows_to_xlsx(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reseñas"

    if not rows:
        wb.save(buffer := io.BytesIO())
        return buffer.getvalue()

    headers = list(rows[0].keys())
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append([row.get(h) for h in headers])

    ws.freeze_panes = "A2"
    for i, header in enumerate(headers, start=1):
        max_len = max([len(header)] + [len(str(r.get(header, ""))) for r in rows[:200]])
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 10), 60)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def paginate(page: int, page_size: int):
    page = max(page, 1)
    page_size = min(max(page_size, 1), 200)
    offset = (page - 1) * page_size
    return page, page_size, offset


def _normalize_mes(value):
    """Acepta 'YYYY-MM', 'MM/YYYY', una fecha de Excel (datetime) o similar,
    y devuelve 'YYYY-MM'. Lanza ValueError si no se puede interpretar."""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m")
    text = str(value).strip()
    m = re.match(r"^(\d{4})-(\d{1,2})$", text)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}"
    m = re.match(r"^(\d{1,2})[/-](\d{4})$", text)
    if m:
        return f"{m.group(2)}-{int(m.group(1)):02d}"
    raise ValueError(f"Mes no reconocido: {value!r} (usa formato YYYY-MM, p.ej. 2026-06)")


def read_transactions_xlsx(file_bytes, default_mes=None):
    """Lee un Excel con columnas 'tienda', 'mes' y 'transacciones' (cabecera
    en cualquier orden/mayúsculas). Si falta la columna 'mes' se usa
    `default_mes` para todas las filas. Devuelve una lista de
    (tienda, mes, transacciones)."""
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"No se pudo leer el archivo Excel: {e}")

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(h).strip().lower() if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        return []

    try:
        tienda_idx = header.index("tienda")
        trans_idx = header.index("transacciones")
    except ValueError:
        raise ValueError("El Excel debe tener columnas 'tienda' y 'transacciones' en la primera fila")
    mes_idx = header.index("mes") if "mes" in header else None
    if mes_idx is None and not default_mes:
        raise ValueError("El Excel debe tener columna 'mes' (o selecciona un mes antes de subirlo)")

    rows = []
    for row in rows_iter:
        if row is None or tienda_idx >= len(row) or trans_idx >= len(row):
            continue
        tienda = row[tienda_idx]
        transacciones = row[trans_idx]
        mes_val = row[mes_idx] if mes_idx is not None and mes_idx < len(row) and row[mes_idx] is not None else default_mes
        if tienda is None or transacciones is None or mes_val is None:
            continue
        try:
            mes = _normalize_mes(mes_val)
            rows.append((str(tienda).strip(), mes, int(transacciones)))
        except (TypeError, ValueError):
            continue
    return rows
